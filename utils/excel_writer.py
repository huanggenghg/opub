# -*- coding: utf-8 -*-
"""
Excel写入工具
用于记录发布成功的视频链接
"""
import os
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None


def get_excel_filename(date_str: str = None, author: str = "黄耿") -> str:
    """
    生成Excel文件名

    Args:
        date_str: 日期字符串，格式为 MMDD，如 "0508"，默认使用当天日期
        author: 作者名称

    Returns:
        str: 文件名，如 "75条自媒体链接-0508-黄耿.xlsx"
    """
    if date_str is None:
        date_str = datetime.now().strftime("%m%d")
    return f"75条自媒体链接-{date_str}-{author}.xlsx"


def get_excel_filepath(output_dir: str = None, date_str: str = None, author: str = "黄耿") -> str:
    """
    获取Excel文件完整路径

    Args:
        output_dir: 输出目录，默认为项目根目录下的 output 文件夹
        date_str: 日期字符串
        author: 作者名称

    Returns:
        str: 文件完整路径
    """
    if output_dir is None:
        project_root = Path(__file__).parent.parent
        output_dir = project_root / "output"

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = get_excel_filename(date_str, author)
    return str(output_dir / filename)


def write_video_link(video_link: str, output_dir: str = None, date_str: str = None, author: str = "黄耿") -> dict:
    """
    写入视频链接到Excel文件（简单模式，只写入链接文本）

    Args:
        video_link: 视频链接
        output_dir: 输出目录
        date_str: 日期字符串
        author: 作者名称

    Returns:
        dict: {"success": bool, "filepath": str, "row": int, "message": str}
    """
    if openpyxl is None:
        raise ImportError("需要安装 openpyxl: pip install openpyxl")

    try:
        filepath = get_excel_filepath(output_dir, date_str, author)

        # 如果文件不存在，创建新文件
        if not os.path.exists(filepath):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "链接"
            next_row = 1  # 新文件从第1行开始
        else:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            # 找到下一个空行
            next_row = ws.max_row + 1

        # 写入链接
        ws.cell(row=next_row, column=1, value=video_link)

        wb.save(filepath)

        return {
            "success": True,
            "filepath": filepath,
            "row": next_row,
            "message": f"已写入第 {next_row} 条链接"
        }
    except Exception as e:
        return {
            "success": False,
            "filepath": "",
            "row": 0,
            "message": str(e)
        }


# 测试代码
if __name__ == "__main__":
    result = write_video_link(
        video_link="https://www.douyin.com/video/7637405747755732270"
    )
    print(f"结果: {result}")
